
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side
import datetime
import argparse
import math
import unicodedata
import os
import sys

# 项目名称：野火鲁班猫310B（LubanCat-310B）
# 项目编号：EBF410686V0R1
# 项目日期：2026年3月19日

def _build_output_file_name(pro_num, pro_date):
    return f"{pro_num}_BOM_{pro_date}.xlsx"

def _display_width(text):
    width = 0
    for char in text:
        width += 2 if unicodedata.east_asian_width(char) in ('F', 'W') else 1
    return width

def _estimate_row_height(cell_value, column_width, line_height=18):
    if cell_value in (None, ''):
        return line_height

    # Excel 列宽和实际字符数不是 1:1，这里做一个近似估算。
    max_chars_per_line = max(1, int(column_width * 1.3))
    wrapped_lines = 0
    for raw_line in str(cell_value).splitlines() or ['']:
        wrapped_lines += max(1, math.ceil(_display_width(raw_line) / max_chars_per_line))
    return wrapped_lines * line_height + 8 # 加上5像素的额外空间

def adjust_excel_format(input_file, output_file, pro_name, pro_num, pro_date=None):
    if pro_date is None:
        pro_date = datetime.datetime.now().strftime("%Y%m%d")

    # Load the existing workbook
    workbook = openpyxl.load_workbook(input_file)
    sheet = workbook.active

    # Insert a new row at the top
    sheet.insert_rows(1)  
    # Insert a new row at the bottom (after the last row)
    sheet.cell(row=sheet.max_row + 1, column=1)
    max_col = sheet.max_column
    max_row = sheet.max_row
    
    print(f"Max rows: {max_row}, Max columns: {max_col}")

    # 合并居中第一行
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=sheet.max_column)
    title_cell = sheet.cell(row=1, column=1)
    # 第一行格式：项目名称 BOM (项目编号 项目日期)
    title_cell.value = f"{pro_name} BOM({pro_num} {pro_date})"  # Set the title text
    title_cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    # 把字体设置为16，字体颜色为红色
    title_cell.font = openpyxl.styles.Font(size=16, color='FF0000')

    # 第二行设置为灰色背景，红色字体，字体大小为11
    for col in range(1, max_col + 1):
        cell = sheet.cell(row=2, column=col)
        cell.fill = openpyxl.styles.PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        cell.font = openpyxl.styles.Font(size=11, color='FF0000')

   # 最后一行字体为红色，字体大小为11，垂直居中
    for col in range(1, max_col + 1):
        cell = sheet.cell(row=max_row, column=col)
        cell.font = openpyxl.styles.Font(color='FF0000', size=11)
        cell.alignment = openpyxl.styles.Alignment(vertical='center')

    # 最后一行：第一列：项目编号，第三列：PCB，第6列：项目编号 日期，第10列：1，其余空白
    sheet.cell(row=max_row, column=1).value = pro_num
    sheet.cell(row=max_row, column=3).value = "PCB"
    sheet.cell(row=max_row, column=6).value = f"{pro_num} {pro_date}"
    sheet.cell(row=max_row, column=10).value = "1"

    # 第一列列宽为19，第二列列宽为45，第三列列宽为18，
    # 第四列列宽为8， 第五列列宽为10，第六列列宽为25，
    # 第七列列宽为45，第八列列宽为25，第九列列宽为25，第十列列宽为10
    column_widths = [19, 45, 18, 8, 10, 25, 45, 25, 25, 10]
    for i, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[get_column_letter(i)].width = width

    # 正文行先统一使用 36 的基础行高，只有换行估算结果超过 36 时才增高。
    for row in range(1, max_row + 1):
        sheet.row_dimensions[row].height = 36

    # 第2、第7列设置为自动换行、垂直居中，并按内容估算正文行高
    for row in range(3, max_row + 1):
        row_height = 36
        for col in [2, 7]:
            cell = sheet.cell(row=row, column=col)
            cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='center')
            estimated_height = _estimate_row_height(cell.value, column_widths[col - 1])
            if estimated_height > 36:
                row_height = max(row_height, estimated_height)
        sheet.row_dimensions[row].height = row_height

    # 最后一列设置垂直居中，水平靠右
    for row in range(1, max_row + 1):
        cell = sheet.cell(row=row, column=max_col)
        cell.alignment = openpyxl.styles.Alignment(horizontal='right', vertical='center')

    # 设置边框为所有框线
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    
    # Apply border to all cells
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = sheet.cell(row=row, column=col)
            cell.border = thin_border

    # 重置滚动位置，防止源文件停在 160 行被原样保留
    view = sheet.sheet_view
    view.topLeftCell = None

    # 冻结前两行
    sheet.freeze_panes = 'A3'
    
    workbook.save(output_file)
    return max_row, max_col


def parse_args():
    parser = argparse.ArgumentParser(description="Format BOM Excel and apply styles.")
    parser.add_argument("input_file", nargs="?", default=None,
                        help="Input Excel file path. Required.")
    parser.add_argument("output_file", nargs="?", default=None,
                        help="Output Excel file path. If omitted, auto-generate with 项目编号_BOM_项目日期.xlsx")
    parser.add_argument("-n", "--name", "--pro-name", dest="pro_name", required=True,
                        help="Project name. Required.")
    parser.add_argument("-N", "--num", "--pro-num", dest="pro_num", required=True,
                        help="Project number. Required.")
    parser.add_argument("-d", "--date", "--pro-date", dest="pro_date", default=None,
                        help="Project date in YYYYMMDD. If omitted, use system date.")
    return parser, parser.parse_args()


def _is_valid_yyyymmdd(date_text):
    if len(date_text) != 8 or not date_text.isdigit():
        return False
    try:
        datetime.datetime.strptime(date_text, "%Y%m%d")
    except ValueError:
        return False
    return True


def validate_args(args):
    if not args.input_file:
        return False
    if not os.path.isfile(args.input_file):
        return False
    if not args.pro_name or not args.pro_name.strip():
        return False
    if not args.pro_num or not args.pro_num.strip():
        return False
    if args.pro_date is not None and not _is_valid_yyyymmdd(args.pro_date):
        return False
    return True


if __name__ == "__main__":
    parser, args = parse_args()

    if not validate_args(args):
        parser.print_help()
        sys.exit(2)

    effective_pro_name = args.pro_name
    effective_pro_num = args.pro_num
    effective_pro_date = args.pro_date or datetime.datetime.now().strftime("%Y%m%d")
    effective_output_file = args.output_file or _build_output_file_name(effective_pro_num, effective_pro_date)

    max_row, max_col = adjust_excel_format(
        args.input_file,
        effective_output_file,
        pro_name=effective_pro_name,
        pro_num=effective_pro_num,
        pro_date=effective_pro_date,
    )

    print("处理完成")
    print(f"输入文件: {args.input_file}")
    print(f"输出文件: {effective_output_file}")
    print(f"项目名称: {effective_pro_name}")
    print(f"项目编号: {effective_pro_num}")
    print(f"项目日期: {effective_pro_date}")
    print(f"工作表规模: {max_row} 行, {max_col} 列")